# -*- coding: utf-8 -*-
import logging

import ddt
import pymysql
from openpyxl import *
import json
from com.zb.control import config


class Datasource():
    def __init__(self):
        pass

    #取excel数据
    @staticmethod
    def get_excel_data(excel_name):
        wb = load_workbook(excel_name)
        ws = wb[wb.sheetnames[0]]
        case_list = []
        title_list = []
        for k in range(1, ws.max_column + 1):  # 用excel的第一列title生成dict的key
            title_list.append(ws.cell(row=1, column=k).value)
        for i in range(2, ws.max_row + 1):  # excel一行一个dict，多行为list
            dict_temp = {}
            for j in range(1, ws.max_column + 1):
                dict_temp[title_list[j - 1]] = ws.cell(row=i, column=j).value
            case_list.append(dict_temp)
        return case_list

    @staticmethod
    def replace_json_kw(template, kw_data):
        json_data = json.loads(template[2])  # 模版list的json转字典
        for key in kw_data:
            if key.find("kw.") >= 0:
                if key.find(".", 3) == -1:  # 一级
                    key_tmp = key.replace("kw.", "")
                    json_data[key_tmp] = kw_data[key]
                elif key.find(".", 3) >= 0:  # 第二，三级
                    if key.find("kw.", key.find(".", 3)) >= 0:  # 第三级
                        pass
                    else:
                        pass  # 第二级
        template[2] = json_data  # 替换原模板中的json
        return template

    # 连db
    @staticmethod
    def get_request_info(interface_name, data_excel=None):
        if data_excel is None:
            data_excel = {}
        try:
            db = pymysql.connect(host=config.DB_HOST, port=config.DB_PORT, user=config.DB_USER, password=config.DB_PWD,
                                 db=config.DB_DATABASE, charset='utf8')
        except Exception, dberr:
            logging.error("connect da error:%s" % dberr)
            return
        cu = db.cursor()
        cu.execute("SELECT a.type,a.url,a.json FROM interface_template a WHERE a.interface_name=%s", interface_name)
        # db.commit()
        result = cu.fetchall()
        try:
            result = list(result[0])
            result = Datasource.replace_json_kw(result, data_excel) #调用替换方法
        except Exception, e:
            logging.error("select errror:%s" % e)
        finally:
            db.close()
            return result


# if __name__ == '__main__':
#     requsts_info = Datasource.get_request_info("productdetail")
#     excel1 = {'ex_amount': 1200L, 'ex_rate': 1L, 'kw.productid': '968732659754864641', 'no': 1L}
#     # Datasource.replace_json_kw(requsts_info, excel1)
