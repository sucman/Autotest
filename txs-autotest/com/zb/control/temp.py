# -*- coding: utf-8 -*-
import logging
import unittest

import ddt as ddt
from openpyxl import *

from com.zb.control.test_report import HtmlReport
from com.zb.testcase.testcase import Pubapi_1


class Datasource():
    def __init__(self):
        pass

    @staticmethod
    def get_excel_data(excel_name):
        # json = {"type":2}
        # return json
        wb = load_workbook(excel_name)
        ws = wb[wb.sheetnames[0]]
        list_dict_name = []
        # for i in range(1, ws.max_column + 1):
        #     list_dict_name.append(ws.cell(row=1, column=i).value)

        CASE_LIST = []
        for i in range(2, ws.max_row + 1):
            list_temp = []
            for j in range(1, ws.max_column + 1):
                list_temp.append(ws.cell(row=i, column=j).value)
            CASE_LIST.append(list_temp)
            # yield list_temp
        return CASE_LIST


@ddt.ddt
class Dome1(unittest.TestCase):
    def setUp(self):
        pass

    def tearDown(self):
        pass

    @ddt.data(*Datasource.get_excel_data("pubapi_banner_all.xlsx"))
    def testcasee(self, data):
        print data[0]
        # assert data[0] == True

    pass


if __name__ == '__main__':
    # a = Datasource.get_excel_data("adtest.xlsx")
    # print a
    suite = unittest.TestLoader().loadTestsFromTestCase(Pubapi_1)
    # unittest.TextTestRunner(verbosity=2).run(suite)
    # print a.next()
    # print a.next()
    # print a.next()


    # suite.addTest(Dome("test_new_welfare"))


    HtmlReport.html_report(suite)
